#!/usr/bin/env python
# coding: utf-8


#imports
from bs4 import BeautifulSoup
import requests
from requests_html import HTMLSession
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from time import sleep


#my wbsites
imoVirtual_apartaments_cascais = 'https://www.imovirtual.com/pt/resultados/comprar/apartamento/lisboa/cascais?limit=72&ownerTypeSingleSelect=ALL&priceMax=240000&roomsNumber=%5BTHREE%2CFOUR%2CFIVE%2CSIX_OR_MORE%5D&by=DEFAULT&direction=DESC&viewType=listing'
imoVirtual_houses_cascais = 'https://www.imovirtual.com/pt/resultados/comprar/moradia/lisboa/cascais?ownerTypeSingleSelect=ALL&roomsNumber=%5BTHREE%2CFOUR%2CFIVE%2CSIX_OR_MORE%5D&priceMax=240000&by=DEFAULT&direction=DESC&viewType=listing'
imoVirtual_apartaments_oeiras = 'https://www.imovirtual.com/pt/resultados/comprar/apartamento/lisboa/oeiras?limit=72&ownerTypeSingleSelect=ALL&priceMax=240000&roomsNumber=%5BTHREE%2CFOUR%2CFIVE%2CSIX_OR_MORE%5D&by=DEFAULT&direction=DESC&viewType=listing'
imoVirtual_houses_oeiras = 'https://www.imovirtual.com/pt/resultados/comprar/moradia/lisboa/oeiras?limit=36&ownerTypeSingleSelect=ALL&priceMax=240000&roomsNumber=%5BTHREE%2CFOUR%2CFIVE%2CSIX_OR_MORE%5D&by=DEFAULT&direction=DESC&viewType=listing'
idealista = 'https://www.idealista.pt/comprar-casas/cascais/com-preco-max_240000,t2,t3,t4-t5/?ordem=atualizado-desc'
remax_cascais = 'https://www.remax.pt/comprar?searchQueryState=%7B%22regionName%22:%22cascais%22,%22businessType%22:1,%22page%22:1,%22regionID%22:%22%22,%22regionType%22:%22%22,%22sort%22:%7B%22fieldToSort%22:%22PublishDate%22,%22order%22:1%7D,%22mapIsOpen%22:false,%22price%22:%7B%22min%22:null,%22max%22:240000%7D,%22mapScroll%22:false,%22rooms%22:2%7D'
remax_oeiras = 'https://www.remax.pt/comprar?searchQueryState=%7B%22regionName%22:%22Oeiras%22,%22businessType%22:1,%22page%22:1,%22regionID%22:%22541%22,%22regionType%22:%22Region2ID%22,%22sort%22:%7B%22fieldToSort%22:%22PublishDate%22,%22order%22:1%7D,%22mapIsOpen%22:false,%22listingClass%22:1,%22price%22:%7B%22min%22:null,%22max%22:240000%7D,%22mapScroll%22:false,%22rooms%22:2,%22listingTypes%22:%5B%5D,%22prn%22:%22Oeiras,%20Lisboa%22,%22regionCoordinates%22:%7B%22latitude%22:38.7170951617666,%22longitude%22:-9.269621200241543%7D,%22regionZoom%22:12%7D'
era_link = 'https://www.era.pt/comprar?ob=1&tp=1,2&lc=11-05,11-10&nqMin=2&nqMax=5&pvMax=250000&page=1&ord=3'
#webiste lists
imoVirtual_list = [imoVirtual_apartaments_cascais, imoVirtual_houses_cascais, imoVirtual_apartaments_oeiras, imoVirtual_houses_oeiras]
remax_list = [remax_cascais, remax_oeiras]


#Function to get information for each website
def imoVirtual (site_url):
    """Provide the Imovirtual link and it will add the houses info to my excel: houses.xlsx
    The excel file must be in the same folder of the code"""
    
    #use headers to mimic a real website
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.107 Safari/537.36',
        'Referer': 'https://www.google.com/',
    }

    #request
    r = requests.get(site_url, headers=headers, verify=False) #using verify=False is not the best practice
    if r.status_code == 200:
        soup = BeautifulSoup(r.text, 'html.parser')
        #get houses information
        sections = soup.find_all('section') # division for all the houses
        #loop for each house
        for section in sections:
            try:
                name = section.find('p', class_="css-u3orbr e1g5xnx10").text
                zone = section.find('p', class_="css-42r2ms eejmx80").text
                house_price = section.find('span', class_='css-2bt9f1 evk7nst0').text
                url = f"https://www.imovirtual.com{section.find('a')['href']}"
                #unpack general info
                dd_elements = section.find_all('dd')
                values = [dd.get_text() for dd in dd_elements]
                bedrooms = values[0]
                area = values[1]
                description = section.find('div', class_="css-1b63dzw e1uq9mc93").text
                #create a list to save my house values
                info_list = [name, zone, house_price, url, bedrooms, area, description]
            except:
                print("Could not get house information.")
            #Load excel file
            workbook = load_workbook(filename='houses.xlsx')
            #Load sheet by index
            sheet = workbook.worksheets[0]
            number_of_rows = sheet.max_row #Used to calculate the last row of the table
            #check if house is in the excel based on the url befor add it
            found = False
            for cell in sheet.iter_rows(min_row=1, max_row=number_of_rows, min_col=4, max_col=4):
                for c in cell:
                    if c.value == url:
                        found = True #found my string (house url) in the excel
                        break
            if found == False: # this means the house url is not in the excel
                #loop to add values to excel
                for numb in range(sheet.max_column): 
                    #add value to a cell
                    sheet.cell(row=number_of_rows + 1, column=numb+1).value = info_list[numb]
                    #Apply color to new rows
                    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    sheet.cell(row=number_of_rows + 1, column=numb+1).fill = yellow_fill   
            #save the file
            workbook.save(filename='houses.xlsx')         
    else:
        print("Not possible to get website.")
    print("FINISHED")


def remaxHouses (my_url):
    # Set up Chrome options for headless mode
    chrome_options = Options()
    chrome_options.add_argument("--headless")  # Ensure GUI is off
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")

    # Automatically manage the WebDriver
    #driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)

    # Run this when you dont need to manage any webdriver update
    driver = webdriver.Chrome(options=chrome_options)

    # Open a URL
    driver.get(my_url)
    sleep(4) # wait for the page to load

    # Get page content
    page_content = driver.page_source
    # parse HTML content with Beautifull Soup
    soup = BeautifulSoup(page_content,'html.parser')

    # Close the browser
    driver.quit()

    #Get house information using Beautifull Soup
    house_div = soup.find_all('div', class_="col-12 col-sm-6 col-md-6 col-lg-4 col-xl-3 result")
    #Loop to get information of all the houses
    for house in house_div:
        name = f'{house.find("li", class_="listing-type").text.strip()} Remax'
        zone = house.find('h2', class_="listing-address").find('span').text.strip()
        price = house.find('p', class_="listing-price").text.strip()
        url = f"https://www.remax.pt{house.find('a')['href']}"
        bedrooms = str("T" + house.find('li', class_="listing-bedroom").text.strip())
        area = house.find('li', class_="listing-area").text.strip()
        description = house.find('span', id="listing-description-tags").text.replace("-", " ")
        #create a list to save my house values
        info_list = [name, zone, price, url, bedrooms, area, description]
        #Load excel file
        workbook = load_workbook(filename='houses.xlsx')
        #Load sheet by index
        sheet = workbook.worksheets[0]
        number_of_rows = sheet.max_row #Used to calculate the last row of the table
        #check if house is in the excel based on the url befor add it
        found = False
        for cell in sheet.iter_rows(min_row=1, max_row=number_of_rows, min_col=4, max_col=4):
            for c in cell:
                if c.value == url:
                    found = True #found my string (house url) in the excel
                    break
        if found == False: # this means the house url is not in the excel
            #loop to add values to excel
            for numb in range(sheet.max_column): 
                #add value to a cell
                sheet.cell(row=number_of_rows + 1, column=numb+1).value = info_list[numb]
                #Apply color to new rows
                yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                sheet.cell(row=number_of_rows + 1, column=numb+1).fill = yellow_fill   
        #save the file
        workbook.save(filename='houses.xlsx')

    print("Finished")


def idealista_houses(my_url):
    #It is not possible to run in headless mode because of captcha

    # Automatically manage the WebDriver
    #driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)

    # Run this when you dont need to manage any webdriver update
    driver = webdriver.Chrome()

    # Open a URL
    driver.get(my_url)
    sleep(20) # wait for the page to load

    # Get page content
    page_content = driver.page_source
    # parse HTML content with Beautifull Soup
    soup = BeautifulSoup(page_content,'html.parser')

    # Close the browser
    driver.quit()

    #Get house information using Beautifull Soup
    house_div = soup.find('div', class_="item-info-container")
    #Loop to get information of all the houses
    for house in house_div:
        name = house.find('a')['title']
        #to get the zone first clean some text spliting and slicing the string and next joining it
        zone_word_list = house.find('a')['title'].split()[3:]
        zone = " ".join(zone_word_list)
        price = house.find('span', class_="item-price h2-simulated").text
        url = f'https://www.idealista.pt{house.find("a")["href"]}'
        #bedroom and area are in the same tag. Use find_all and use the slice of the list
        bedrooms = house.find_all('span', class_="item-detail")[0].text.strip()
        area = house.find_all('span', class_="item-detail")[1].text.strip().replace(" área bruta", "")
        description = house.find('div', class_="item-description description").text.strip()
        #create a list to save my house values
        info_list = [name, zone, price, url, bedrooms, area, description]
        #Load excel file
        workbook = load_workbook(filename='houses.xlsx')
        #Load sheet by index
        sheet = workbook.worksheets[0]
        number_of_rows = sheet.max_row #Used to calculate the last row of the table
        #check if house is in the excel based on the url befor add it
        found = False
        for cell in sheet.iter_rows(min_row=1, max_row=number_of_rows, min_col=4, max_col=4):
            for c in cell:
                if c.value == url:
                    found = True #found my string (house url) in the excel
                    break
        if found == False: # this means the house url is not in the excel
            #loop to add values to excel
            for numb in range(sheet.max_column): 
                #add value to a cell
                sheet.cell(row=number_of_rows + 1, column=numb+1).value = info_list[numb]
                #Apply color to new rows
                yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                sheet.cell(row=number_of_rows + 1, column=numb+1).fill = yellow_fill   
        #save the file
        workbook.save(filename='houses.xlsx')

    print("Finished")


def era_houses(my_url):
    # Set up Chrome options for headless mode
    chrome_options = Options()
    chrome_options.add_argument("--headless")  # Ensure GUI is off
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")

    # Automatically manage the WebDriver
    #driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)

    # Run this when you dont need to manage any webdriver update
    driver = webdriver.Chrome(options=chrome_options)

    # Open a URL
    driver.get(my_url)
    sleep(5) # wait for the page to load

    # Get page content
    page_content = driver.page_source
    # parse HTML content with Beautifull Soup
    soup = BeautifulSoup(page_content,'html.parser')

    # Close the browser
    driver.quit()

    #Get house information using Beautifull Soup
    house_div = soup.find_all(class_="content p-3")
    #Loop to get information of all the houses
    for house in house_div:
        name = f'{house.find("p", class_="property-type d-block mb-1").text} ERA'
        zone = house.find('div', class_="col-12 location").text
        price = house.find('p', class_="price-value").text
        url = house.find('a')['href']
        bedrooms = f'T{house.find_all("span", class_="d-inline-flex")[0].text}'
        area = house.find_all("span", class_="d-inline-flex")[3].text
        description = "No description"
        #create a list to save my house values
        info_list = [name, zone, price, url, bedrooms, area, description]
        #Load excel file
        workbook = load_workbook(filename='houses.xlsx')
        #Load sheet by index
        sheet = workbook.worksheets[0]
        number_of_rows = sheet.max_row #Used to calculate the last row of the table
        #check if house is in the excel based on the url befor add it
        found = False
        for cell in sheet.iter_rows(min_row=1, max_row=number_of_rows, min_col=4, max_col=4):
            for c in cell:
                if c.value == url:
                    found = True #found my string (house url) in the excel
                    break
        if found == False: # this means the house url is not in the excel
            #loop to add values to excel
            for numb in range(sheet.max_column): 
                #add value to a cell
                sheet.cell(row=number_of_rows + 1, column=numb+1).value = info_list[numb]
                #Apply color to new rows
                yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                sheet.cell(row=number_of_rows + 1, column=numb+1).fill = yellow_fill   
        #save the file
        workbook.save(filename='houses.xlsx')

    print("Finished")


#Run code for each website
for link in imoVirtual_list:
    imoVirtual(link)

for link in remax_list:
    remaxHouses(link)

try: #this website block frequently the driver
    idealista_houses(idealista)
except:
    print("Website block driver.")

era_houses(era_link)
print("End of code.")